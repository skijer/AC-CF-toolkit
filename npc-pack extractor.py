# Import necessary libraries
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from tqdm import tqdm  # Import tqdm for progress bar

def read_bin_file(file_path):
    # Open file as a binary
    with open(file_path, 'rb') as bin_file:
        return bin_file.read()

def read_hex(data, offset):
    # Convert the bytes at the specified offset to an integer
    byte_hex = data[offset:offset+1]
    byte = int(byte_hex.hex(), 16)
    return byte

def feature_hex_transform(data, offset, df):
    # Convert the bytes at the specified offset to an item on the item list  and divided by 4 natural spacing
    byte_hex = data[offset:offset+2]
    byte = int(byte_hex.hex(), 16) - 36864
    byte_div = byte / 4
    
    # Use the DataFrame passed as an argument and search for the row
    row = df[df['DEC ID'] == byte_div]
    
    # Check if the row was found and return the value from the column 'English' Change to a language you want
    if not row.empty:
        return str(row['English'].values[0])
    else:
        return None


def get_high_nibble(byte):
    return byte >> 4
def get_low_nibble(byte):
    return byte & 0x0F

def parse_block(data, block_number, df):
    block_data = {}
    offset=00
    # House information
    house = {}
    furnitures = ["space","Master model number", "Default shirt", "Default Floor", "Default Wall", "Default Parasol", "Item 01", "Item 02", "Item 03", "Item 04", "Item 05", "Item 06", "Item 07", "Item 09", "Item 10", "Item 11", "K.K. Song","Unknown"]
    for ftr in furnitures:
        if ftr == "space":
            value = data[offset:offset+1].hex().upper()
        elif ftr == "Master model number":
            value = str(read_hex(data, offset))
        elif ftr == "Unknown":
            value = data[offset:offset+2].hex().upper()
        else:
            value = feature_hex_transform(data, offset, df)
            offset += 1
        house[ftr] = value
        offset += 1
    block_data["House"] = house
    
    # Different languages names
    names = {}
    offset = 0x22
    languages = ["Japanese", "English", "Spanish America", "Spanish", "French", "Italian", "German", "Korean"]
    for language in languages:
        name = data[offset:offset + 18]
        name_str = name.decode('utf-16-be', errors='ignore').strip('\x00')
        names[language] = name_str
        offset += 18
    block_data["Names"] = names

    # Catchphrases on different languages:
    catchphrases = {}
    offset = 0xB2
    languages = ["Catch-Japanese", "Catch-English US", "Catch-Spanish America", "Catch-French Canada", "Catch-English", "Catch-Spanish", "Catch-French", "Catch-Italian", "Catch-German", "Catch-Korean"]
    for language in languages:
        phrase = data[offset:offset + 22]
        phrase_str = phrase.decode('utf-16-be', errors='ignore').strip('\x00')
        catchphrases[language] = phrase_str
        offset += 22
    block_data["Catchphrases"] = catchphrases

    # Villager stats variables:
    stats = {}
    stats_name = ["Specie", "Month of birth", "Day of birth", "Unknown-Stat", "Favorite clothing", "Less favorite clothing", "Favorite furniture color", "Favorite furniture series", "Personality", "Favorite furniture styles", "Starting villager"]
    for stat in stats_name:
        if stat == "Specie":
            stat_characteristics = ["cat", "elephant", "sheep", "bear", "dog", "squirrel", "rabbit", "duck", "hip", "wolf", "mouse", "pig", "chicken", "bull", "cow", "bird", "frog", "alligator", "goat", "tiger", "anteater", "koala", "horse", "octopus", "lion", "bear cub", "rhinoceros", "gorilla", "ostrich", "kangaroo", "eagle", "penguin", "monkey"]
            value = stat_characteristics[read_hex(data, offset)]
        elif stat == "Month of birth" or stat == "Day of birth":
            value = str(read_hex(data, offset))
        elif stat == "Unknown-Stat":
            value = data[offset:offset+1].hex().upper()
        elif stat == "Favorite clothing" or stat == "Less favorite clothing":
            stat_characteristics = ["cute", "cool", "subtle", "gaudy", "strange", "funky", "refined", "fresh", "stylish", "striking"]
            value = stat_characteristics[read_hex(data, offset)]
        elif stat == "Favorite furniture color":
            stat_characteristics = ["", "yellow", "red", "orange", "green", "blue", "white", "black", "purple", "brown", "pink", "gray", "colorful", "aqua", "beige"]
            value = stat_characteristics[read_hex(data, offset)]
        elif stat == "Favorite furniture series":
            stat_characteristics = ["Exotic series", "Lovely series", "Classic series", "Ranch series", "Cabana series", "Blue series", "Modern series", "Regal series", "Green series", "Cabin series", "Kiddies series", "Robo series"]
            value = stat_characteristics[read_hex(data, offset)]
        elif stat == "Personality":
            stat_characteristics = ["Lazy","Jock","Cranky","Normal","Peppy","Snooty"]
            temp = get_high_nibble(read_hex(data, offset))
            value = stat_characteristics[temp]
            offset -= 1
        elif stat== "Favorite furniture styles":
            stat_characteristics = ["","","","","","Playful and Retro", "Dignified and Retro","","","Playful and Trendy","Dignified and Trendy"]
            temp = get_low_nibble(read_hex(data, offset))
            value = stat_characteristics[temp]
        elif stat == "Starting villager":
            value = "Yes" if data[offset:offset+1].hex()=='80' else "No"

        stats[stat] = value
        offset += 1
    block_data["Stats"] = stats

    return block_data

def save_to_excel(parsed_data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Villager Data"
    
    headers = ["Block Number"] + list(parsed_data[0]["House"].keys()) + list(parsed_data[0]["Names"].keys()) + list(parsed_data[0]["Catchphrases"].keys()) + list(parsed_data[0]["Stats"].keys())
    ws.append(headers)

    for block_number, block in enumerate(parsed_data, start=1):
        row = [block_number]
        row.extend(block["House"].values())
        row.extend(block["Names"].values())
        row.extend(block["Catchphrases"].values())
        row.extend(block["Stats"].values())
        ws.append(row)

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_path)

def process_bin_file(input_path, output_path):
    binary_data = read_bin_file(input_path)
    df = pd.read_excel("src/aurum's-item-list.xlsx", engine='openpyxl')
    block_size = 408  # villager memory size

    results = []
    offset = 0x20 # Header
    block_number = 1

    # Initialize tqdm for progress bar
    pbar = tqdm(total=len(binary_data) // block_size, desc="Processing binary file", unit="block")

    while offset + block_size <= len(binary_data):
        block_data = binary_data[offset:offset + block_size]
        results.append(parse_block(block_data, block_number, df))
        offset += block_size
        block_number += 1
        pbar.update(1)  # Update progress bar    
        # Finalize tqdm progress bar
    pbar.close()

    save_to_excel(results, output_path)

# Initializer
process_bin_file('pack.bin', 'pack.acdat')
